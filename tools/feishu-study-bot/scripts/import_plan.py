#!/usr/bin/env python3
"""
学习计划导入脚本
支持从Excel文件导入到数据库
"""
import sys
import json
import re
from datetime import datetime, timedelta
from pathlib import Path

# 添加项目根目录到路径
sys.path.insert(0, str(Path(__file__).parent.parent))

from models.database import init_db, SessionLocal, WeeklyPlan, Deliverable, Milestone


def parse_excel_plan(filepath: str) -> dict:
    """解析Excel格式的学习计划"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("请先安装 openpyxl: pip install openpyxl")
        sys.exit(1)
    
    wb = load_workbook(filepath)
    
    plan = {
        "weeks": [],
        "milestones": []
    }
    
    # 解析Weekly Plan sheet (sheet2)
    if "Weekly Plan" in wb.sheetnames:
        ws = wb["Weekly Plan"]
        rows = list(ws.iter_rows(values_only=True))
        
        # 跳过表头
        for row in rows[1:]:
            if not row[0]:
                continue
            
            # 解析日期范围
            date_range = str(row[1]) if len(row) > 1 else ""
            start_date, end_date = parse_date_range(date_range)
            
            # 解析交付物
            deliverables = []
            if len(row) > 3 and row[3]:  # 关键交付物列
                deliv_text = str(row[3])
                for line in deliv_text.split('\n'):
                    if line.strip():
                        deliverables.append({
                            "name": line.strip()[:200],
                            "description": "",
                            "acceptance_criteria": str(row[5]) if len(row) > 5 else ""
                        })
            
            week = {
                "week_num": extract_week_num(str(row[0])),
                "start_date": start_date,
                "end_date": end_date,
                "stage": str(row[2]) if len(row) > 2 else "",
                "theme": str(row[4]) if len(row) > 4 else "",
                "focus": str(row[4]) if len(row) > 4 else "",
                "main_goal": "vLLM 交付模板",
                "deep_dive": str(row[8]) if len(row) > 8 else "",
                "trl_task": str(row[9]) if len(row) > 9 else "",
                "slime_task": str(row[10]) if len(row) > 10 else "",
                "openrlhf_task": str(row[11]) if len(row) > 11 else "",
                "risks": str(row[13]) if len(row) > 13 else "",
                "deliverables": deliverables[:5]  # 最多5个交付物
            }
            plan["weeks"].append(week)
    
    # 解析Milestones sheet (sheet4)
    if "Milestones" in wb.sheetnames:
        ws = wb["Milestones"]
        rows = list(ws.iter_rows(values_only=True))
        
        for row in rows[1:]:
            if not row[0]:
                continue
            
            milestone = {
                "name": str(row[0]),
                "target_week": int(row[1]) if len(row) > 1 and row[1] else 0,
                "description": str(row[2]) if len(row) > 2 else "",
                "completion_criteria": str(row[3]) if len(row) > 3 else ""
            }
            plan["milestones"].append(milestone)
    
    return plan


def parse_date_range(date_str: str) -> tuple:
    """解析日期范围"""
    # 格式: 2026-02-01 ~ 2026-02-07
    match = re.search(r'(\d{4}-\d{2}-\d{2}).*~(\d{4}-\d{2}-\d{2})', date_str)
    if match:
        return match.group(1), match.group(2)
    return "", ""


def extract_week_num(text: str) -> int:
    """提取周次数字"""
    match = re.search(r'W(\d+)', text, re.IGNORECASE)
    if match:
        return int(match.group(1))
    # 尝试直接解析数字
    match = re.search(r'\d+', text)
    if match:
        return int(match.group())
    return 0


def import_to_database(plan: dict):
    """导入到数据库"""
    db = SessionLocal()
    try:
        # 清空现有数据
        db.query(Deliverable).delete()
        db.query(WeeklyPlan).delete()
        db.query(Milestone).delete()
        db.commit()
        
        # 导入周计划
        for week_data in plan["weeks"]:
            if week_data["week_num"] == 0:
                continue
            
            week = WeeklyPlan(
                week_num=week_data["week_num"],
                start_date=week_data["start_date"],
                end_date=week_data["end_date"],
                stage=week_data["stage"],
                theme=week_data["theme"],
                focus=week_data["focus"],
                main_goal=week_data["main_goal"],
                deep_dive=week_data["deep_dive"],
                trl_task=week_data["trl_task"],
                slime_task=week_data["slime_task"],
                openrlhf_task=week_data["openrlhf_task"],
                risks=week_data["risks"]
            )
            db.add(week)
            db.flush()  # 获取ID
            
            # 导入交付物
            for deliv_data in week_data["deliverables"]:
                deliverable = Deliverable(
                    week_id=week.id,
                    name=deliv_data["name"][:200],
                    description=deliv_data["description"],
                    acceptance_criteria=deliv_data["acceptance_criteria"]
                )
                db.add(deliverable)
        
        # 导入里程碑
        for ms_data in plan["milestones"]:
            milestone = Milestone(
                name=ms_data["name"],
                target_week=ms_data["target_week"],
                description=ms_data["description"],
                completion_criteria=ms_data["completion_criteria"]
            )
            db.add(milestone)
        
        db.commit()
        print(f"✅ 成功导入 {len(plan['weeks'])} 周计划和 {len(plan['milestones'])} 个里程碑")
        
    except Exception as e:
        db.rollback()
        print(f"❌ 导入失败: {e}")
        raise
    finally:
        db.close()


def main():
    if len(sys.argv) < 2:
        print("用法: python import_plan.py <excel文件路径>")
        print("示例: python import_plan.py ../周计划_20260201-20260701.xlsx")
        sys.exit(1)
    
    filepath = sys.argv[1]
    
    print(f"📖 正在读取: {filepath}")
    plan = parse_excel_plan(filepath)
    
    print(f"📊 解析结果:")
    print(f"   - 周计划: {len(plan['weeks'])} 周")
    print(f"   - 里程碑: {len(plan['milestones'])} 个")
    
    # 保存为JSON备份
    json_path = filepath.replace('.xlsx', '_parsed.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(plan, f, ensure_ascii=False, indent=2)
    print(f"💾 已保存解析结果到: {json_path}")
    
    # 导入数据库
    print("📝 正在导入数据库...")
    init_db()
    import_to_database(plan)
    
    print("✨ 完成!")


if __name__ == "__main__":
    main()
